from io import BytesIO
from httpx import AsyncClient
from openpyxl import load_workbook


class TestImportExport:
    async def _create_sample_data(self, client: AsyncClient, headers: dict):
        for i in range(3):
            await client.post("/api/initiatives", json={
                "q": f"Q{i+1}",
                "account": f"Account {i}",
                "unit": f"Unit {i}",
                "lpr": f"LPR {i}",
                "action": f"Action {i}",
                "kpi": f"KPI {i}",
                "priority": "high" if i % 2 == 0 else "critical",
                "status": "active" if i < 2 else "done",
                "owner": f"Owner {i}",
                "potential": float(i * 10),
                "next_date": "",
                "comment": "",
            }, headers=headers)

    async def test_export_empty(self, client: AsyncClient, auth_headers: dict):
        resp = await client.get("/api/export", headers=auth_headers)
        assert resp.status_code == 200
        content_type = resp.headers["content-type"]
        assert "spreadsheetml.sheet" in content_type
        assert "filename" in resp.headers.get("content-disposition", "")

    async def test_export_with_data(self, client: AsyncClient, auth_headers: dict):
        await self._create_sample_data(client, auth_headers)
        resp = await client.get("/api/export", headers=auth_headers)
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content))
        assert "Дорожная карта" in wb.sheetnames
        assert "KPI Дашборд" in wb.sheetnames
        assert "Трекер контактов" in wb.sheetnames

        ws = wb["Дорожная карта"]
        header_row = [ws.cell(row=1, column=c).value for c in range(1, 13)]
        assert "Квартал" in header_row
        assert "Аккаунт" in header_row

        data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(row))
        assert data_rows >= 3
        wb.close()

    async def test_import_invalid_file(self, client: AsyncClient, auth_headers: dict):
        resp = await client.post("/api/import", files={"file": ("test.txt", b"not an xlsx", "text/plain")}, headers=auth_headers)
        assert resp.status_code == 400

    async def test_import_roundtrip(self, client: AsyncClient, auth_headers: dict):
        """Import exported data back - should work"""
        await self._create_sample_data(client, auth_headers)
        export_resp = await client.get("/api/export", headers=auth_headers)
        assert export_resp.status_code == 200

        import_resp = await client.post(
            "/api/import",
            files={"file": ("test.xlsx", export_resp.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
        assert import_resp.status_code == 200
        data = import_resp.json()
        assert data["imported"] == 3

    async def test_import_clears_old_data(self, client: AsyncClient, auth_headers: dict):
        await client.post("/api/initiatives", json={"q": "Q1", "account": "Old", "unit": "x"}, headers=auth_headers)
        assert len((await client.get("/api/initiatives", headers=auth_headers)).json()) == 1

        from app.services.excel_service import generate_workbook
        buf = generate_workbook([{"q": "Q1", "account": "New", "unit": "y", "lpr": "", "action": "",
                                   "kpi": "", "priority": "high", "status": "pending", "owner": "",
                                   "potential": 0, "next_date": "", "comment": ""}], [])
        resp = await client.post("/api/import", files={"file": ("test.xlsx", buf.getvalue(),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                                 headers=auth_headers)
        assert resp.status_code == 200
        items = (await client.get("/api/initiatives", headers=auth_headers)).json()
        assert len(items) == 1
        assert items[0]["account"] == "New"
