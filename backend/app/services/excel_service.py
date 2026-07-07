from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.constants import Q_META, HEADERS, CONTACT_HEADERS, STATUS_LABELS_EMOJI, InitiativeStatus, Priority

HEADER_FILL = PatternFill("solid", fgColor="1A2C4E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="DDE3EF"))


def parse_workbook(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    ws_name = next(
        (n for n in wb.sheetnames if "карт" in n.lower()),
        wb.sheetnames[0],
    )
    initiatives = _parse_roadmap_sheet(wb[ws_name])
    contacts = _parse_contacts_sheet(wb)

    wb.close()
    return initiatives, contacts


def _parse_roadmap_sheet(ws) -> list[dict]:
    initiatives = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [v if v is not None else "" for v in row]
        if len(values) < 12:
            continue
        q_raw = str(values[0]).strip()
        if not q_raw or q_raw.startswith("▶"):
            continue
        account = str(values[1]).strip() if len(values) > 1 else ""
        unit = str(values[2]).strip() if len(values) > 2 else ""
        if not account and not unit:
            continue

        initiatives.append({
            "q": _norm_quarter(q_raw),
            "account": account,
            "unit": unit,
            "lpr": str(values[3]).strip() if len(values) > 3 else "",
            "action": str(values[4]).strip() if len(values) > 4 else "",
            "kpi": str(values[5]).strip() if len(values) > 5 else "",
            "priority": _norm_priority(str(values[6])) if len(values) > 6 else Priority.HIGH,
            "status": _norm_status(str(values[7])) if len(values) > 7 else InitiativeStatus.PENDING,
            "owner": str(values[8]).strip() if len(values) > 8 else "",
            "potential": _parse_float(values[9]) if len(values) > 9 else 0,
            "next_date": str(values[10]).strip() if len(values) > 10 else "",
            "comment": str(values[11]).strip() if len(values) > 11 else "",
        })
    return initiatives


def _parse_contacts_sheet(wb) -> list[dict]:
    c_name = next((n for n in wb.sheetnames if "контакт" in n.lower()), None)
    if not c_name:
        return []

    contacts = []
    for row in wb[c_name].iter_rows(min_row=2, values_only=True):
        values = [v if v is not None else "" for v in row]
        if not values[0] or "ТРЕКЕР" in str(values[0]).upper():
            continue
        contacts.append({
            "account": str(values[0]).strip(),
            "unit": str(values[1]).strip() if len(values) > 1 else "",
            "name": str(values[2]).strip() if len(values) > 2 else "",
            "email": str(values[3]).strip() if len(values) > 3 else "",
            "phone": str(values[4]).strip() if len(values) > 4 else "",
            "last_date": str(values[5]).strip() if len(values) > 5 else "",
            "topic": str(values[6]).strip() if len(values) > 6 else "",
            "next_step": str(values[7]).strip() if len(values) > 7 else "",
        })
    return contacts


def _norm_quarter(v: str) -> str:
    for q in ["1", "2", "3", "4"]:
        if q in v.lower():
            return f"Q{q}"
    return "Q1"


def _norm_priority(v: str) -> Priority:
    return Priority.CRITICAL if "крит" in v.lower() else Priority.HIGH


def _norm_status(v: str) -> InitiativeStatus:
    s = v.lower()
    if "работ" in s:
        return InitiativeStatus.ACTIVE
    if "ожид" in s or "жд" in s:
        return InitiativeStatus.WAITING
    if "выполн" in s:
        return InitiativeStatus.DONE
    if "риск" in s:
        return InitiativeStatus.RISK
    return InitiativeStatus.PENDING


def _parse_float(v) -> float:
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def generate_workbook(initiatives: list[dict], contacts: list[dict]) -> BytesIO:
    wb = Workbook()

    _build_roadmap_sheet(wb, initiatives)
    _build_kpi_sheet(wb, initiatives)
    _build_contacts_sheet(wb, contacts)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf


def _style_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_roadmap_sheet(wb, initiatives: list[dict]):
    ws = wb.active
    ws.title = "Дорожная карта"
    _style_header(ws, HEADERS)

    row_num = 2
    last_q = None
    for row in initiatives:
        if row["q"] != last_q:
            last_q = row["q"]
            label = Q_META[row["q"]]["label"].upper()
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=12)
            cell = ws.cell(row=row_num, column=1, value=f"▶  {label}")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="B0C4DE")
            row_num += 1

        priority_label = "🔴 Критический" if row["priority"] == "critical" else "🟡 Высокий"
        status_label = STATUS_LABELS_EMOJI.get(row["status"], row["status"])

        values = [
            Q_META[row["q"]]["label"],
            row["account"], row["unit"], row["lpr"], row["action"], row["kpi"],
            priority_label, status_label,
            row["owner"], row["potential"] or 0, row["next_date"], row["comment"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = THIN_BORDER
        row_num += 1

    widths = [10] + [22] * 11
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_kpi_sheet(wb, initiatives: list[dict]):
    ws = wb.create_sheet("KPI Дашборд")
    total = len(initiatives)
    done = sum(1 for r in initiatives if r["status"] == "done")
    pct = round(done / total * 100) if total else 0

    kpi_data = [
        ["Показатель", "Значение"],
        ["Всего инициатив", total],
        ["Выполнено", done],
        ["% выполнения", f"{pct}%"],
        ["Критических", sum(1 for r in initiatives if r["priority"] == "critical")],
        ["Под риском", sum(1 for r in initiatives if r["status"] == "risk")],
        ["Общий потенциал, ₽ млн", sum(r.get("potential", 0) for r in initiatives)],
        ["Закрытый потенциал, ₽ млн", sum(r.get("potential", 0) for r in initiatives if r["status"] == "done")],
    ]
    for r, row_data in enumerate(kpi_data, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def _build_contacts_sheet(wb, contacts: list[dict]):
    ws = wb.create_sheet("Трекер контактов")
    _style_header(ws, CONTACT_HEADERS)

    fields = ["account", "unit", "name", "email", "phone", "last_date", "topic", "next_step"]
    for r, c in enumerate(contacts, 2):
        for ci, field in enumerate(fields, 1):
            cell = ws.cell(row=r, column=ci, value=c.get(field, ""))
            cell.border = THIN_BORDER

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22
